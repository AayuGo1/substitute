"""Download and load the engineering monitoring workbook from GitHub.

This module is responsible for one thing only: turning a configured
GitHub location into a loaded :class:`openpyxl.workbook.workbook.Workbook`
object. It does not parse worksheet contents, does not validate workbook
structure, does not compute KPIs, and contains no other business logic —
those responsibilities belong to ``data/validator.py`` and later modules.

Configuration contract
-----------------------
This module reads its settings from ``config/github.py``, which is
expected to expose:

``GitHubConfig``
    A configuration object with the following attributes:

    - ``owner: str`` -- the GitHub account or organization that owns the
      repository.
    - ``repo: str`` -- the repository name.
    - ``branch: str`` -- the branch to read the workbook from.
    - ``file_path: str`` -- the path to the workbook file within the
      repository.
    - ``raw_url: Optional[str]`` -- an explicit raw file URL to use
      instead of building one from ``owner``/``repo``/``branch``/
      ``file_path``. Leave as ``None`` to build the URL automatically.
    - ``timeout_seconds: float`` -- how long to wait for the download
      before giving up.
    - ``cache_ttl_seconds: Optional[float]`` -- how long a downloaded
      workbook may be served from cache before being re-downloaded,
      resolved once when this module is imported. ``None`` means the
      cached workbook never expires on its own (use
      :func:`clear_workbook_cache` to force a refresh).

``get_github_config() -> GitHubConfig``
    A factory function that returns the active configuration.

This module never hardcodes an owner, repository, branch, or file path;
every one of those values is supplied by ``config/github.py``, so pointing
the dashboard at a different repository or file requires no code change
here.
"""

from __future__ import annotations

import zipfile
from dataclasses import dataclass
from io import BytesIO
from typing import Optional
from urllib.parse import urlparse

import requests
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.workbook import Workbook

from config.github import GitHubConfig, get_github_config

_RAW_GITHUB_HOST = "raw.githubusercontent.com"
_ALLOWED_URL_SCHEMES = ("http", "https")


# --------------------------------------------------------------------------
# Custom exceptions
# --------------------------------------------------------------------------


class WorkbookLoaderError(Exception):
    """Base exception for every workbook loading failure."""


class GitHubConfigurationError(WorkbookLoaderError):
    """Raised when the GitHub configuration is missing or unusable."""


class InvalidWorkbookURLError(WorkbookLoaderError):
    """Raised when the resolved workbook URL is missing or malformed."""


class WorkbookDownloadError(WorkbookLoaderError):
    """Raised when the workbook cannot be downloaded from GitHub."""


class WorkbookDownloadTimeoutError(WorkbookDownloadError):
    """Raised when downloading the workbook exceeds the configured timeout."""


class CorruptedWorkbookDownloadError(WorkbookLoaderError):
    """Raised when the downloaded content is not a valid Excel workbook."""


# --------------------------------------------------------------------------
# Configuration resolution
# --------------------------------------------------------------------------


def _resolve_active_config(config: Optional[GitHubConfig]) -> GitHubConfig:
    """Resolve the GitHub configuration to use for a download.

    Args:
        config: An explicit configuration to use, or ``None`` to fetch the
            active configuration from ``config/github.py``.

    Returns:
        The configuration to use for this download.

    Raises:
        GitHubConfigurationError: If no configuration could be obtained.
    """
    if config is not None:
        return config
    try:
        return get_github_config()
    except Exception as exc:  # noqa: BLE001 - any config failure is fatal
        raise GitHubConfigurationError(
            f"Failed to load GitHub configuration from config/github.py: {exc}"
        ) from exc


def build_raw_url(config: GitHubConfig) -> str:
    """Build a raw GitHub file URL from a configuration's repository settings.

    Args:
        config: The configuration to build a URL for. If ``config.raw_url``
            is set, it is returned unchanged; otherwise a raw URL is built
            from ``owner``, ``repo``, ``branch``, and ``file_path``.

    Returns:
        The resolved raw file URL.

    Raises:
        GitHubConfigurationError: If the configuration has neither a
            ``raw_url`` nor a complete set of repository settings.
    """
    if getattr(config, "raw_url", None):
        return config.raw_url  # type: ignore[return-value]

    required_fields = ("owner", "repo", "branch", "file_path")
    missing_fields = [
        field_name
        for field_name in required_fields
        if not getattr(config, field_name, None)
    ]
    if missing_fields:
        raise GitHubConfigurationError(
            "GitHub configuration is incomplete; cannot build a raw file "
            f"URL without: {', '.join(missing_fields)}."
        )

    return (
        f"https://{_RAW_GITHUB_HOST}/{config.owner}/{config.repo}/"
        f"{config.branch}/{config.file_path}"
    )


# --------------------------------------------------------------------------
# URL validation
# --------------------------------------------------------------------------


def validate_workbook_url(url: str) -> None:
    """Validate that a workbook URL is well-formed before attempting a download.

    Args:
        url: The URL to validate.

    Raises:
        InvalidWorkbookURLError: If the URL is empty, has no scheme, has no
            host, or uses a scheme other than ``http``/``https``.
    """
    if not url or not url.strip():
        raise InvalidWorkbookURLError("Workbook URL is empty.")

    parsed = urlparse(url)
    if parsed.scheme not in _ALLOWED_URL_SCHEMES:
        raise InvalidWorkbookURLError(
            f"Workbook URL '{url}' has an unsupported scheme "
            f"'{parsed.scheme}'; expected one of {_ALLOWED_URL_SCHEMES}."
        )
    if not parsed.netloc:
        raise InvalidWorkbookURLError(f"Workbook URL '{url}' has no host.")


# --------------------------------------------------------------------------
# Download and parsing
# --------------------------------------------------------------------------


def download_workbook_bytes(url: str, timeout_seconds: float) -> bytes:
    """Download raw file content from a URL, translating failures clearly.

    Args:
        url: The URL to download from. Must already be validated.
        timeout_seconds: How long to wait for the request before giving up.

    Returns:
        The downloaded response body as raw bytes.

    Raises:
        WorkbookDownloadTimeoutError: If the request exceeds
            ``timeout_seconds``.
        InvalidWorkbookURLError: If the URL is rejected by the HTTP client
            as malformed.
        WorkbookDownloadError: If the request fails for any other network
            or HTTP reason, or returns no content.
    """
    try:
        response = requests.get(url, timeout=timeout_seconds)
        response.raise_for_status()
    except requests.exceptions.Timeout as exc:
        raise WorkbookDownloadTimeoutError(
            f"Timed out after {timeout_seconds}s downloading workbook from '{url}'."
        ) from exc
    except (
        requests.exceptions.MissingSchema,
        requests.exceptions.InvalidSchema,
        requests.exceptions.InvalidURL,
    ) as exc:
        raise InvalidWorkbookURLError(f"Workbook URL '{url}' is invalid: {exc}") from exc
    except requests.exceptions.HTTPError as exc:
        status_code = getattr(exc.response, "status_code", "unknown")
        raise WorkbookDownloadError(
            f"GitHub returned HTTP {status_code} when downloading '{url}'."
        ) from exc
    except requests.exceptions.ConnectionError as exc:
        raise WorkbookDownloadError(
            f"A network error occurred while downloading workbook from '{url}': {exc}"
        ) from exc
    except requests.exceptions.RequestException as exc:
        raise WorkbookDownloadError(
            f"Failed to download workbook from '{url}': {exc}"
        ) from exc

    if not response.content:
        raise WorkbookDownloadError(f"Downloaded workbook from '{url}' is empty.")
    return response.content


def parse_workbook_bytes(content: bytes) -> Workbook:
    """Load an openpyxl Workbook from raw downloaded bytes.

    Args:
        content: The raw bytes downloaded from GitHub.

    Returns:
        The loaded workbook, with formulas resolved to their
        last-calculated values.

    Raises:
        CorruptedWorkbookDownloadError: If the bytes do not represent a
            valid, structurally sound Excel workbook.
    """
    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except zipfile.BadZipFile as exc:
        raise CorruptedWorkbookDownloadError(
            "Downloaded file is not a valid Excel workbook (corrupted archive)."
        ) from exc
    except InvalidFileException as exc:
        raise CorruptedWorkbookDownloadError(
            f"Downloaded file has an invalid or unsupported format: {exc}"
        ) from exc
    except KeyError as exc:
        raise CorruptedWorkbookDownloadError(
            f"Downloaded workbook is missing required internal structure: {exc}"
        ) from exc
    except Exception as exc:  # noqa: BLE001 - any other parse failure is fatal
        raise CorruptedWorkbookDownloadError(
            f"Failed to open downloaded workbook: {exc}"
        ) from exc

    if not workbook.sheetnames:
        raise CorruptedWorkbookDownloadError(
            "Downloaded workbook contains no worksheets."
        )
    return workbook


# --------------------------------------------------------------------------
# Cached download pipeline
# --------------------------------------------------------------------------


@dataclass(frozen=True)
class _CacheKey:
    """An explicit, hashable key describing one downloadable workbook source.

    Kept separate from :class:`GitHubConfig` so the Streamlit cache is
    keyed only on the values that actually affect the downloaded content,
    regardless of what other fields the configuration object may carry.

    Attributes:
        url: The resolved raw file URL.
        timeout_seconds: The download timeout used for this source.
    """

    url: str
    timeout_seconds: float


def _download_and_parse(cache_key: _CacheKey) -> Workbook:
    """Download and parse a workbook for a given cache key.

    Args:
        cache_key: The resolved URL and timeout to download with.

    Returns:
        The loaded workbook.
    """
    content = download_workbook_bytes(cache_key.url, cache_key.timeout_seconds)
    return parse_workbook_bytes(content)


def _resolve_default_cache_ttl_seconds() -> Optional[float]:
    """Best-effort lookup of the configured cache TTL at import time.

    Returns:
        The ``cache_ttl_seconds`` value from the active GitHub
        configuration, or ``None`` if it is unavailable (for example
        during testing, before configuration is fully set up), meaning
        the cache will not expire on its own.
    """
    try:
        return get_github_config().cache_ttl_seconds
    except Exception:  # noqa: BLE001 - fall back to no expiry if unavailable
        return None


_DEFAULT_CACHE_TTL_SECONDS: Optional[float] = _resolve_default_cache_ttl_seconds()


@st.cache_resource(
    show_spinner="Downloading engineering monitoring workbook from GitHub...",
    ttl=_DEFAULT_CACHE_TTL_SECONDS,
)
def _cached_download_and_parse(cache_key: _CacheKey) -> Workbook:
    """Run the download-and-parse pipeline behind Streamlit's resource cache.

    Decorated once at module import time, as Streamlit's cache decorators
    only provide real caching when applied to a stable function object;
    re-applying the decorator on every call would create a fresh, empty
    cache each time and defeat caching entirely.

    Args:
        cache_key: The resolved URL and timeout to download with. Acts as
            the cache key, so a different source URL or timeout is always
            downloaded fresh.

    Returns:
        The loaded workbook, served from cache when available.
    """
    return _download_and_parse(cache_key)


def clear_workbook_cache() -> None:
    """Clear the cached workbook so the next call re-downloads from GitHub.

    Intended to back a manual "reload data" action in the UI layer.
    """
    _cached_download_and_parse.clear()


# --------------------------------------------------------------------------
# Public entry point
# --------------------------------------------------------------------------


def load_workbook_from_github(config: Optional[GitHubConfig] = None) -> Workbook:
    """Download and load the engineering monitoring workbook from GitHub.

    This is the only function other modules should call to obtain a
    workbook. It resolves configuration, builds and validates the source
    URL, and serves the result from Streamlit's resource cache so repeated
    calls within the configured TTL do not trigger repeated downloads.

    Args:
        config: An explicit :class:`GitHubConfig` to use instead of the
            active configuration returned by
            ``config.github.get_github_config``. Primarily useful for
            testing against an alternate source.

    Returns:
        The loaded workbook, containing only raw cell data. No parsing,
        validation, or calculation has been performed on its contents.

    Raises:
        GitHubConfigurationError: If no usable configuration is available.
        InvalidWorkbookURLError: If the resolved source URL is missing or
            malformed.
        WorkbookDownloadTimeoutError: If the download exceeds the
            configured timeout.
        WorkbookDownloadError: If the download fails for any other network
            or HTTP reason.
        CorruptedWorkbookDownloadError: If the downloaded content is not a
            valid Excel workbook.
    """
    active_config = _resolve_active_config(config)
    url = build_raw_url(active_config)
    validate_workbook_url(url)

    timeout_seconds = getattr(active_config, "timeout_seconds", 15.0)
    cache_key = _CacheKey(url=url, timeout_seconds=timeout_seconds)

    return _cached_download_and_parse(cache_key)

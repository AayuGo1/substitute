"""Structural validation for Excel workbooks used by the monitoring dashboard.

This module verifies that an uploaded workbook is loadable, structurally
sound, and internally consistent enough to be safely parsed later. It does
not parse business data and does not perform any calculations.

All validation logic is discovery-based: header rows, column positions,
date columns, and numeric columns are all identified by inspecting the
actual contents of the workbook at validation time. Nothing here assumes a
fixed row number, column letter, department name, section name, or header
position, so the same code continues to work when the monthly workbook is
replaced with another file that shares the same logical structure.
"""

from __future__ import annotations

import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple, Union

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_MISSING_VALUE_MARKERS: Tuple[str, ...] = (
    "",
    "-",
    "--",
    "---",
    "n/a",
    "na",
    "null",
    "none",
)


# --------------------------------------------------------------------------
# Custom exceptions
# --------------------------------------------------------------------------


class WorkbookValidationError(Exception):
    """Base exception for every workbook validation failure."""


class WorkbookLoadError(WorkbookValidationError):
    """Raised when the workbook file cannot be opened at all."""


class CorruptedWorkbookError(WorkbookValidationError):
    """Raised when the workbook file is unreadable or structurally invalid."""


class MissingWorksheetError(WorkbookValidationError):
    """Raised when one or more required worksheets are absent."""


class EmptySheetError(WorkbookValidationError):
    """Raised when a worksheet contains no data at all."""


class HeaderValidationError(WorkbookValidationError):
    """Raised when a worksheet has no readable header row."""


class DuplicateColumnError(WorkbookValidationError):
    """Raised when a worksheet's header row contains duplicate column names."""


class InvalidDateColumnError(WorkbookValidationError):
    """Raised when a detected date-like column contains non-date values."""


class InvalidNumericColumnError(WorkbookValidationError):
    """Raised when a detected numeric-like column contains non-numeric values."""


# --------------------------------------------------------------------------
# Data containers
# --------------------------------------------------------------------------


@dataclass(frozen=True)
class ValidationIssue:
    """A single, non-fatal data quality problem found during validation.

    Attributes:
        sheet_name: Name of the worksheet the issue was found in.
        category: Short machine-readable label for the kind of issue.
        message: Human-readable description of the issue.
        column: Optional 1-based column index the issue relates to.
    """

    sheet_name: str
    category: str
    message: str
    column: Optional[int] = None

    def __str__(self) -> str:
        """Render the issue as a single descriptive line."""
        location = f" (column={self.column})" if self.column is not None else ""
        return f"[{self.sheet_name}] {self.category}: {self.message}{location}"


@dataclass(frozen=True)
class SheetValidationReport:
    """Summary of the structure discovered in a single worksheet.

    Attributes:
        sheet_name: Name of the validated worksheet.
        header_row_index: 1-based row number identified as the header row.
        headers: Header labels discovered on the header row, in column order.
        data_row_count: Number of rows below the header row that exist.
        column_count: Number of columns with a readable header label.
    """

    sheet_name: str
    header_row_index: int
    headers: Tuple[str, ...]
    data_row_count: int
    column_count: int


@dataclass
class WorkbookValidationReport:
    """Aggregated outcome of validating an entire workbook.

    Attributes:
        sheet_reports: Per-sheet structural summaries, keyed by sheet name.
        warnings: Non-fatal data quality issues collected across all sheets.
    """

    sheet_reports: Dict[str, SheetValidationReport] = field(default_factory=dict)
    warnings: List[ValidationIssue] = field(default_factory=list)

    @property
    def has_warnings(self) -> bool:
        """Return True if any non-fatal issues were collected."""
        return bool(self.warnings)


# --------------------------------------------------------------------------
# Loading and sheet-presence validation
# --------------------------------------------------------------------------


def load_workbook_safely(file_path: Union[str, Path]) -> Workbook:
    """Load an Excel workbook, translating low-level failures into clear errors.

    Args:
        file_path: Path to the ``.xlsx`` workbook file to load.

    Returns:
        The loaded openpyxl Workbook, with formulas resolved to their
        last-calculated values.

    Raises:
        WorkbookLoadError: If the file does not exist or cannot be opened
            for a reason other than corruption.
        CorruptedWorkbookError: If the file exists but is not a valid,
            structurally sound Excel workbook.
    """
    path = Path(file_path)
    if not path.exists():
        raise WorkbookLoadError(f"Workbook file does not exist: {path}")
    if not path.is_file():
        raise WorkbookLoadError(f"Workbook path is not a file: {path}")

    try:
        workbook = load_workbook(filename=str(path), data_only=True)
    except zipfile.BadZipFile as exc:
        raise CorruptedWorkbookError(
            f"Workbook is not a valid Excel file (corrupted archive): {path}"
        ) from exc
    except InvalidFileException as exc:
        raise CorruptedWorkbookError(
            f"Workbook has an invalid or unsupported file format: {path}"
        ) from exc
    except KeyError as exc:
        raise CorruptedWorkbookError(
            f"Workbook '{path}' is missing required internal structure: {exc}"
        ) from exc
    except Exception as exc:  # noqa: BLE001 - any other load failure is fatal
        raise WorkbookLoadError(f"Failed to load workbook '{path}': {exc}") from exc

    if not workbook.sheetnames:
        raise CorruptedWorkbookError(f"Workbook '{path}' contains no worksheets.")
    return workbook


def validate_required_worksheets(
    workbook: Workbook, required_sheet_names: Sequence[str]
) -> None:
    """Ensure every required worksheet name exists in the workbook.

    Args:
        workbook: The loaded workbook to check.
        required_sheet_names: Worksheet names that must be present. An
            empty sequence means no particular sheet is required.

    Raises:
        MissingWorksheetError: If any required worksheet is absent.
    """
    if not required_sheet_names:
        return
    available = set(workbook.sheetnames)
    missing = [name for name in required_sheet_names if name not in available]
    if missing:
        raise MissingWorksheetError(
            f"Missing required worksheet(s): {', '.join(missing)}. "
            f"Available worksheets: {', '.join(workbook.sheetnames)}."
        )


# --------------------------------------------------------------------------
# Sheet emptiness validation
# --------------------------------------------------------------------------


def is_sheet_empty(sheet: Worksheet) -> bool:
    """Determine whether a worksheet contains no data in any cell.

    Args:
        sheet: The worksheet to inspect.

    Returns:
        True if every cell in the worksheet is empty, False otherwise.
    """
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                return False
    return True


def validate_sheet_not_empty(sheet: Worksheet) -> None:
    """Raise if a worksheet contains no data at all.

    Args:
        sheet: The worksheet to validate.

    Raises:
        EmptySheetError: If the worksheet is completely empty.
    """
    if is_sheet_empty(sheet):
        raise EmptySheetError(f"Worksheet '{sheet.title}' is completely empty.")


# --------------------------------------------------------------------------
# Header discovery and validation
# --------------------------------------------------------------------------


def detect_header_row(sheet: Worksheet, max_rows_to_scan: int = 15) -> int:
    """Dynamically identify which row in a worksheet is the header row.

    The header row is assumed to be the most text-dense row near the top
    of the sheet: real header rows list one label per column, while rows
    above them (titles, grouped section labels) or below them (data) are
    comparatively sparse in text.

    Args:
        sheet: The worksheet to inspect.
        max_rows_to_scan: How many rows from the top to consider as
            header-row candidates.

    Returns:
        The 1-based row index identified as the header row.

    Raises:
        HeaderValidationError: If no row within the scan range contains a
            readable set of header labels.
    """
    scan_limit = min(max_rows_to_scan, sheet.max_row or 0)
    best_row: Optional[int] = None
    best_score = 0
    for row_index in range(1, scan_limit + 1):
        score = sum(
            1
            for cell in sheet[row_index]
            if isinstance(cell.value, str) and cell.value.strip()
        )
        if score > best_score:
            best_score = score
            best_row = row_index

    minimum_required_labels = 1
    if best_row is None or best_score < minimum_required_labels:
        raise HeaderValidationError(
            f"Worksheet '{sheet.title}' has no readable header row within "
            f"the first {scan_limit} row(s)."
        )
    return best_row


def extract_headers(sheet: Worksheet, header_row: int) -> Dict[int, str]:
    """Extract non-empty header labels from a row, keyed by column index.

    Args:
        sheet: The worksheet to read from.
        header_row: The 1-based row index containing header labels.

    Returns:
        A mapping of 1-based column index to trimmed header label, for
        every column in the row that has a readable text label.

    Raises:
        HeaderValidationError: If the row has no readable text labels.
    """
    headers: Dict[int, str] = {}
    for cell in sheet[header_row]:
        if isinstance(cell.value, str) and cell.value.strip():
            headers[cell.column] = cell.value.strip()
    if not headers:
        raise HeaderValidationError(
            f"Worksheet '{sheet.title}' row {header_row} contains no readable headers."
        )
    return headers


def find_duplicate_headers(headers: Iterable[str]) -> List[str]:
    """Find header labels that occur more than once, ignoring case and spacing.

    Args:
        headers: Header labels to check.

    Returns:
        A sorted list of the (lower-cased, trimmed) labels that repeat.
    """
    occurrence_counts: Dict[str, int] = {}
    for header in headers:
        normalized = header.strip().lower()
        occurrence_counts[normalized] = occurrence_counts.get(normalized, 0) + 1
    return sorted(name for name, count in occurrence_counts.items() if count > 1)


def validate_no_duplicate_headers(headers: Iterable[str], sheet_name: str) -> None:
    """Raise if a worksheet's headers contain duplicate labels.

    Args:
        headers: Header labels discovered for the worksheet.
        sheet_name: Name of the worksheet the headers came from, used in
            the error message.

    Raises:
        DuplicateColumnError: If any header label repeats.
    """
    duplicates = find_duplicate_headers(headers)
    if duplicates:
        raise DuplicateColumnError(
            f"Worksheet '{sheet_name}' has duplicate column headers: "
            f"{', '.join(duplicates)}."
        )


# --------------------------------------------------------------------------
# Value-level helpers
# --------------------------------------------------------------------------


def is_missing_value(
    value: Any, missing_markers: Sequence[str] = DEFAULT_MISSING_VALUE_MARKERS
) -> bool:
    """Determine whether a cell value represents a missing/blank entry.

    Args:
        value: The raw cell value.
        missing_markers: Lower-cased string tokens (besides ``None``) that
            should be treated as missing data.

    Returns:
        True if the value should be treated as missing.
    """
    if value is None:
        return True
    if isinstance(value, str) and value.strip().lower() in missing_markers:
        return True
    return False


def get_column_values(
    sheet: Worksheet, column_index: int, start_row: int, end_row: Optional[int] = None
) -> List[Tuple[int, Any]]:
    """Read all values in a column between two rows, inclusive.

    Args:
        sheet: The worksheet to read from.
        column_index: 1-based column index to read.
        start_row: 1-based row index to start reading from.
        end_row: 1-based row index to stop reading at. Defaults to the
            worksheet's last used row.

    Returns:
        A list of (row_index, value) tuples for the requested range.
    """
    last_row = end_row if end_row is not None else (sheet.max_row or start_row)
    return [
        (row_index, sheet.cell(row=row_index, column=column_index).value)
        for row_index in range(start_row, last_row + 1)
    ]


def classify_column_type(
    values: Sequence[Any],
    missing_markers: Sequence[str] = DEFAULT_MISSING_VALUE_MARKERS,
    majority_threshold: float = 0.8,
) -> Optional[str]:
    """Infer whether a column is date-like or numeric-like from its values.

    Classification is based purely on the actual contents of the column,
    never on its position or header name, so it adapts automatically to
    any workbook layout.

    Args:
        values: Raw cell values for the column (including missing entries).
        missing_markers: String tokens treated as missing data.
        majority_threshold: Minimum fraction of non-missing values that
            must share a type for the column to be classified as that type.

    Returns:
        ``"date"``, ``"numeric"``, or ``None`` if the column has no data or
        no single type dominates it.
    """
    present = [value for value in values if not is_missing_value(value, missing_markers)]
    if not present:
        return None

    date_count = sum(1 for value in present if isinstance(value, (datetime, date)))
    numeric_count = sum(
        1
        for value in present
        if isinstance(value, (int, float)) and not isinstance(value, bool)
    )

    if date_count / len(present) >= majority_threshold:
        return "date"
    if numeric_count / len(present) >= majority_threshold:
        return "numeric"
    return None


def find_invalid_date_values(
    values: Sequence[Tuple[int, Any]],
    missing_markers: Sequence[str] = DEFAULT_MISSING_VALUE_MARKERS,
) -> List[Tuple[int, Any]]:
    """Find entries in a date-like column that are not valid dates.

    Args:
        values: (row_index, value) pairs for the column.
        missing_markers: String tokens treated as missing data.

    Returns:
        The (row_index, value) pairs that are neither missing nor a date.
    """
    return [
        (row_index, value)
        for row_index, value in values
        if not is_missing_value(value, missing_markers)
        and not isinstance(value, (datetime, date))
    ]


def find_invalid_numeric_values(
    values: Sequence[Tuple[int, Any]],
    missing_markers: Sequence[str] = DEFAULT_MISSING_VALUE_MARKERS,
) -> List[Tuple[int, Any]]:
    """Find entries in a numeric-like column that are not valid numbers.

    Args:
        values: (row_index, value) pairs for the column.
        missing_markers: String tokens treated as missing data.

    Returns:
        The (row_index, value) pairs that are neither missing nor numeric.
    """
    return [
        (row_index, value)
        for row_index, value in values
        if not is_missing_value(value, missing_markers)
        and (isinstance(value, bool) or not isinstance(value, (int, float)))
    ]


def validate_date_column(
    sheet_name: str,
    column_index: int,
    values: Sequence[Tuple[int, Any]],
    missing_markers: Sequence[str] = DEFAULT_MISSING_VALUE_MARKERS,
) -> None:
    """Raise if a date-like column contains any non-date, non-missing value.

    Args:
        sheet_name: Name of the worksheet the column belongs to.
        column_index: 1-based column index being validated.
        values: (row_index, value) pairs for the column.
        missing_markers: String tokens treated as missing data.

    Raises:
        InvalidDateColumnError: If any offending value is found.
    """
    invalid_values = find_invalid_date_values(values, missing_markers)
    if invalid_values:
        row_index, value = invalid_values[0]
        raise InvalidDateColumnError(
            f"Worksheet '{sheet_name}' column {column_index} contains "
            f"{len(invalid_values)} invalid date value(s); first offender "
            f"at row {row_index}: {value!r}."
        )


def validate_numeric_column(
    sheet_name: str,
    column_index: int,
    values: Sequence[Tuple[int, Any]],
    missing_markers: Sequence[str] = DEFAULT_MISSING_VALUE_MARKERS,
) -> None:
    """Raise if a numeric-like column contains any non-numeric, non-missing value.

    Args:
        sheet_name: Name of the worksheet the column belongs to.
        column_index: 1-based column index being validated.
        values: (row_index, value) pairs for the column.
        missing_markers: String tokens treated as missing data.

    Raises:
        InvalidNumericColumnError: If any offending value is found.
    """
    invalid_values = find_invalid_numeric_values(values, missing_markers)
    if invalid_values:
        row_index, value = invalid_values[0]
        raise InvalidNumericColumnError(
            f"Worksheet '{sheet_name}' column {column_index} contains "
            f"{len(invalid_values)} invalid numeric value(s); first offender "
            f"at row {row_index}: {value!r}."
        )


# --------------------------------------------------------------------------
# Orchestration
# --------------------------------------------------------------------------


class WorkbookValidator:
    """Validates the structural integrity of a monitoring-dashboard workbook.

    The validator discovers header rows, column layouts, and column data
    types by inspecting the workbook's actual contents rather than assuming
    any fixed row, column, department, or section layout. This allows the
    same validator to keep working when the monthly workbook is replaced
    with another file that has the same logical structure but different
    row counts, column counts, or ordering.
    """

    def __init__(
        self,
        required_sheet_names: Optional[Sequence[str]] = None,
        header_scan_row_limit: int = 15,
        missing_value_markers: Sequence[str] = DEFAULT_MISSING_VALUE_MARKERS,
    ) -> None:
        """Configure the validator.

        Args:
            required_sheet_names: Worksheet names that must exist in any
                workbook this validator is asked to validate. Pass ``None``
                or an empty sequence to validate whatever sheets are present
                without requiring specific names.
            header_scan_row_limit: How many rows from the top of each sheet
                to consider as header-row candidates.
            missing_value_markers: String tokens (besides ``None``) that
                represent missing data in this workbook family.
        """
        self._required_sheet_names: Tuple[str, ...] = tuple(required_sheet_names or ())
        self._header_scan_row_limit = header_scan_row_limit
        self._missing_value_markers: Tuple[str, ...] = tuple(
            marker.strip().lower() for marker in missing_value_markers
        )

    def validate(
        self, file_path: Union[str, Path], strict: bool = False
    ) -> WorkbookValidationReport:
        """Run the full validation pipeline against a workbook file.

        Structural problems that make a sheet impossible to read at all
        (an unreadable file, missing worksheets, an empty sheet, or a sheet
        with no readable header row) always raise a descriptive exception,
        since parsing cannot safely proceed without them. Data quality
        problems that a sheet can still be read despite (duplicate header
        labels, unexpected values inside a date-like or numeric-like
        column) are collected as warnings in the returned report instead of
        aborting validation.

        Args:
            file_path: Path to the workbook file to validate.
            strict: If True, also raise an exception when any data quality
                warnings are found instead of returning them in the report.

        Returns:
            A report describing the structure discovered in each validated
            worksheet, plus any non-fatal data quality warnings.

        Raises:
            WorkbookLoadError: If the file cannot be opened.
            CorruptedWorkbookError: If the file is not a valid workbook.
            MissingWorksheetError: If a required worksheet is absent.
            EmptySheetError: If a validated worksheet has no data.
            HeaderValidationError: If a validated worksheet has no
                readable header row.
            DuplicateColumnError: If a validated worksheet's headers
                contain duplicates.
            WorkbookValidationError: If ``strict`` is True and any data
                quality warnings were found.
        """
        workbook = load_workbook_safely(file_path)
        validate_required_worksheets(workbook, self._required_sheet_names)

        sheet_names_to_validate = self._required_sheet_names or tuple(workbook.sheetnames)
        report = WorkbookValidationReport()

        for sheet_name in sheet_names_to_validate:
            sheet = workbook[sheet_name]
            sheet_report, sheet_warnings = self._validate_sheet(sheet)
            report.sheet_reports[sheet_name] = sheet_report
            report.warnings.extend(sheet_warnings)

        if strict and report.warnings:
            raise WorkbookValidationError(
                f"Workbook failed strict validation with {len(report.warnings)} "
                "issue(s): " + "; ".join(str(issue) for issue in report.warnings)
            )
        return report

    def _validate_sheet(
        self, sheet: Worksheet
    ) -> Tuple[SheetValidationReport, List[ValidationIssue]]:
        """Validate a single worksheet's structure and sample its data quality.

        Args:
            sheet: The worksheet to validate.

        Returns:
            A tuple of the sheet's structural report and any non-fatal
            data quality issues found in its columns.
        """
        validate_sheet_not_empty(sheet)
        header_row = detect_header_row(sheet, self._header_scan_row_limit)
        headers_by_column = extract_headers(sheet, header_row)

        sheet_report = SheetValidationReport(
            sheet_name=sheet.title,
            header_row_index=header_row,
            headers=tuple(headers_by_column.values()),
            data_row_count=max((sheet.max_row or header_row) - header_row, 0),
            column_count=len(headers_by_column),
        )

        issues: List[ValidationIssue] = self._collect_duplicate_header_issues(
            sheet.title, headers_by_column.values()
        )
        data_start_row = header_row + 1
        if data_start_row > (sheet.max_row or header_row):
            return sheet_report, issues

        for column_index in headers_by_column:
            values = get_column_values(sheet, column_index, data_start_row)
            column_type = classify_column_type(
                [value for _, value in values], self._missing_value_markers
            )
            if column_type == "date":
                issues.extend(
                    self._collect_date_issues(sheet.title, column_index, values)
                )
            elif column_type == "numeric":
                issues.extend(
                    self._collect_numeric_issues(sheet.title, column_index, values)
                )

        return sheet_report, issues

    def _collect_duplicate_header_issues(
        self, sheet_name: str, headers: Iterable[str]
    ) -> List[ValidationIssue]:
        """Convert duplicate header labels into a single validation issue.

        Args:
            sheet_name: Name of the worksheet the headers belong to.
            headers: Header labels discovered for the worksheet.

        Returns:
            A list with a single summarized issue if any labels repeat,
            otherwise an empty list.
        """
        duplicates = find_duplicate_headers(headers)
        if not duplicates:
            return []
        return [
            ValidationIssue(
                sheet_name=sheet_name,
                category="duplicate_column_headers",
                message=(
                    f"{len(duplicates)} column header label(s) repeat more than "
                    f"once: {', '.join(duplicates)}."
                ),
            )
        ]

    def _collect_date_issues(

        self, sheet_name: str, column_index: int, values: Sequence[Tuple[int, Any]]
    ) -> List[ValidationIssue]:
        """Convert invalid entries in a date-like column into validation issues.

        Args:
            sheet_name: Name of the worksheet the column belongs to.
            column_index: 1-based column index being checked.
            values: (row_index, value) pairs for the column.

        Returns:
            A list with a single summarized issue if any invalid values
            were found, otherwise an empty list.
        """
        invalid_values = find_invalid_date_values(values, self._missing_value_markers)
        if not invalid_values:
            return []
        offending_row, offending_value = invalid_values[0]
        return [
            ValidationIssue(
                sheet_name=sheet_name,
                category="invalid_date_value",
                message=(
                    f"{len(invalid_values)} value(s) in this date column are not "
                    f"valid dates; first offender at row {offending_row}: "
                    f"{offending_value!r}."
                ),
                column=column_index,
            )
        ]

    def _collect_numeric_issues(
        self, sheet_name: str, column_index: int, values: Sequence[Tuple[int, Any]]
    ) -> List[ValidationIssue]:
        """Convert invalid entries in a numeric-like column into validation issues.

        Args:
            sheet_name: Name of the worksheet the column belongs to.
            column_index: 1-based column index being checked.
            values: (row_index, value) pairs for the column.

        Returns:
            A list with a single summarized issue if any invalid values
            were found, otherwise an empty list.
        """
        invalid_values = find_invalid_numeric_values(values, self._missing_value_markers)
        if not invalid_values:
            return []
        offending_row, offending_value = invalid_values[0]
        return [
            ValidationIssue(
                sheet_name=sheet_name,
                category="invalid_numeric_value",
                message=(
                    f"{len(invalid_values)} value(s) in this numeric column are not "
                    f"valid numbers; first offender at row {offending_row}: "
                    f"{offending_value!r}."
                ),
                column=column_index,
            )
        ]

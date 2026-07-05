"""Parsing of an already-loaded workbook into the dashboard's typed models.

This module turns an arbitrary, workbook-defined layout of departments,
meters, and daily/hourly readings into the strongly typed
:class:`~models.workbook.Workbook` object graph consumed by the rest of
the dashboard. It contains no domain knowledge: it does not know the name
of any sheet, section, sub-section, department, or metric ahead of time.
Every one of those is *discovered* by inspecting the shape of the data at
runtime (which rows are text, which are dates, which are numbers, which
cells are populated versus blank).

Architectural boundaries
-------------------------
This module deliberately does **not** touch disk. Loading the ``.xlsx``
file (via ``openpyxl.load_workbook`` or otherwise) is the responsibility
of the Loader/Repository layer; that layer hands this module an
already-open ``openpyxl`` workbook object. Keeping I/O out of the parser
keeps it trivially testable against in-memory workbooks and lets the
Loader/Repository own caching, retries, and file-system concerns on its
own.

Three responsibilities are kept separate, each with its own class:

* :class:`WorkbookStructureDiscoverer` — **discovery**. Inspects a raw
  workbook's cells and infers, per sheet, where the header row, grouping
  rows, unit row, data block, and timestamp column live. Produces a
  :class:`WorkbookStructure`, which is a plain, serializable-shaped
  description of that layout containing no model objects.
* :class:`SheetModelBuilder` — **model construction**. Given a sheet's
  raw grid of values and its already-discovered :class:`SheetLayout`,
  builds the :class:`~models.metric.Metric`, :class:`~models.section.Section`,
  and :class:`~models.section.SubSection` model objects.
* :class:`ParserService` — **parsing** (orchestration). Given an
  already-loaded workbook, either reuses a :class:`WorkbookStructure`
  supplied by the caller (typically produced once by a validator during
  the load/validate step) or asks the discoverer to compute one, then
  drives the model builder over every sheet and assembles the final
  :class:`~models.workbook.Workbook`.

Because discovery is split out from both parsing and model construction,
a workbook's structure can be discovered once (for example as part of
validation) and reused across multiple parses without re-inspecting the
worksheet cells.

No row numbers, column letters, section names, department names, or
metric names are ever hardcoded; only structural heuristics are used.

This module contains parsing logic only: no Streamlit, no Plotly, no KPI
math, no chart construction, and no file I/O.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from numbers import Number
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook
from openpyxl.worksheet.worksheet import Worksheet

from models.metric import Metric, MetricDataPoint
from models.section import DateRange, Section, SubSection
from models.workbook import Workbook, WorkbookMetadata, ValidationStatus


# ----------------------------------------------------------------------
# Shared, stateless cell/value helpers
# ----------------------------------------------------------------------
# These are plain functions (not methods) because both the discoverer and
# the model builder need identical notions of "populated", "data-like",
# and "text-like" cells, and neither should own the other's utilities.

def _read_grid(worksheet: Worksheet) -> List[List[Any]]:
    """Materializes a worksheet's values into a plain 2D list.

    Reading into a plain grid once keeps every downstream helper simple,
    side-effect free, and independent of openpyxl's iteration quirks.

    Args:
        worksheet: The worksheet to read.

    Returns:
        A list of rows, each a list of cell values, 0-indexed internally
        but interpreted as 1-based when addressed elsewhere in this
        module (row 1 is ``grid[0]``).
    """
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def _cell(grid: List[List[Any]], row: int, col: int) -> Any:
    """Fetches a 1-based (row, column) cell, tolerating short rows."""
    if row < 1 or row > len(grid):
        return None
    row_values = grid[row - 1]
    if col < 1 or col > len(row_values):
        return None
    return row_values[col - 1]


def _row_width(grid: List[List[Any]]) -> int:
    """Returns the widest row in the grid, in number of columns."""
    return max((len(row) for row in grid), default=0)


def _is_populated(value: Any) -> bool:
    """Whether a cell should be treated as non-blank."""
    if value is None:
        return False
    if isinstance(value, str) and not value.strip():
        return False
    return True


def _is_data_like(value: Any) -> bool:
    """Whether a cell looks like a data value rather than a label.

    Numbers and dates are considered data-like; ordinary text is not.
    """
    if value is None:
        return False
    if isinstance(value, bool):
        return False
    if isinstance(value, (Number, datetime)):
        return True
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return False
        if _parse_datetime(stripped) is not None:
            return True
        try:
            float(stripped.replace(",", ""))
            return True
        except ValueError:
            return False
    return False


def _is_text_like(value: Any) -> bool:
    """Whether a cell looks like a text label rather than a value."""
    return _is_populated(value) and not _is_data_like(value)


def _parse_datetime(value: Any) -> Optional[datetime]:
    """Attempts to interpret a value as a timestamp.

    Args:
        value: The candidate value, typically a cell's contents.

    Returns:
        A parsed :class:`datetime`, or ``None`` if the value is not a
        recognizable date or time string.
    """
    if isinstance(value, datetime):
        return value
    if not isinstance(value, str):
        return None
    text = value.strip()
    if not text:
        return None
    candidate_formats = (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y/%m/%d %H.%M",
        "%Y/%m/%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%m/%d/%Y",
    )
    for fmt in candidate_formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _coerce_numeric(value: Any) -> Optional[float]:
    """Converts a cell value to a float when meaningfully possible."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, Number):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip().replace(",", "")
        if not stripped:
            return None
        try:
            return float(stripped)
        except ValueError:
            return None
    return None


def _slugify(label: str) -> str:
    """Normalizes a header label into a stable, internal identifier."""
    text = label.strip().lower()
    text = re.sub(r"[^\w]+", "_", text)
    return text.strip("_") or "unnamed_metric"


# ----------------------------------------------------------------------
# Structure representation (discovery output / parsing input)
# ----------------------------------------------------------------------

@dataclass(frozen=True)
class SheetLayout:
    """The structural coordinates discovered for a single worksheet.

    All indices are 1-based, matching openpyxl's own row/column
    numbering, and every field is discovered at runtime; none are ever
    assumed ahead of time.

    Attributes:
        header_row: The row holding column labels (metric names).
        grouping_rows: Rows above the header holding section / sub-section
            labels, ordered from outermost to innermost.
        unit_row: The row directly below the header holding unit labels,
            if one was detected.
        data_start_row: The first row of the contiguous data block.
        data_end_row: The last row of the contiguous data block.
        timestamp_column: The column index carrying timestamp values, if
            one was detected.
        data_columns: The column indices that have a populated header
            label and therefore represent a metric.
    """

    header_row: int
    grouping_rows: Tuple[int, ...]
    unit_row: Optional[int]
    data_start_row: int
    data_end_row: int
    timestamp_column: Optional[int]
    data_columns: Tuple[int, ...]


@dataclass(frozen=True)
class WorkbookStructure:
    """The discovered structural layout for every sheet in a workbook.

    This is a plain description of *where* things live in a workbook —
    it holds no model objects (no :class:`Metric`, :class:`Section`, and
    so on) and no cell values, so it can be computed once (for example by
    a validator during load/validate) and safely reused across multiple
    parses of the same in-memory workbook without re-inspecting cells.

    Attributes:
        sheet_layouts: Mapping of sheet name to its discovered
            :class:`SheetLayout`. A sheet whose structure could not be
            confidently discovered is simply absent from this mapping.
        unparsable_sheets: Mapping of sheet name to a human-readable
            reason, for sheets whose structure could not be discovered.
    """

    sheet_layouts: Dict[str, SheetLayout] = field(default_factory=dict)
    unparsable_sheets: Dict[str, str] = field(default_factory=dict)


# ----------------------------------------------------------------------
# Discovery
# ----------------------------------------------------------------------

class WorkbookStructureDiscoverer:
    """Discovers a workbook's structure without building any model objects.

    The discoverer never names or assumes any particular sheet, section,
    department, or metric; everything is discovered heuristically from
    the shape of the populated cells at runtime. It operates purely on
    an already-open ``openpyxl`` workbook (or a raw grid of values) and
    performs no file I/O of its own.

    Attributes:
        max_header_search_rows: How many leading rows of a sheet to
            inspect when looking for the header row.
        min_data_fill_ratio: The minimum fraction of a row's cells (within
            the candidate data columns) that must be populated for the
            row to be considered part of the data block.
    """

    def __init__(
        self,
        max_header_search_rows: int = 30,
        min_data_fill_ratio: float = 0.3,
    ) -> None:
        """Initializes the discoverer with tunable structural thresholds.

        Args:
            max_header_search_rows: How many leading rows to scan for a
                header row candidate.
            min_data_fill_ratio: Minimum populated-cell ratio for a row to
                still count as part of the data block.
        """
        self.max_header_search_rows = max_header_search_rows
        self.min_data_fill_ratio = min_data_fill_ratio

    def discover(self, raw_workbook: OpenpyxlWorkbook) -> WorkbookStructure:
        """Discovers the structure of every sheet in an open workbook.

        Args:
            raw_workbook: An already-loaded ``openpyxl`` workbook. This
                method never opens or reads a file itself.

        Returns:
            The :class:`WorkbookStructure` describing every sheet whose
            layout could be discovered, plus reasons for any that
            couldn't.
        """
        structure = WorkbookStructure()
        for sheet_name in raw_workbook.sheetnames:
            worksheet = raw_workbook[sheet_name]
            grid = _read_grid(worksheet)
            layout, reason = self.discover_sheet(grid, sheet_name)
            if layout is not None:
                structure.sheet_layouts[sheet_name] = layout
            else:
                structure.unparsable_sheets[sheet_name] = reason or (
                    "sheet structure could not be discovered"
                )
        return structure

    def discover_sheet(
        self, grid: List[List[Any]], sheet_name: str
    ) -> Tuple[Optional[SheetLayout], Optional[str]]:
        """Discovers the layout of a single sheet's already-read grid.

        Args:
            grid: The worksheet's values as a 2D list.
            sheet_name: The name of the sheet, used only in messages.

        Returns:
            A ``(layout, reason)`` tuple. ``layout`` is ``None`` and
            ``reason`` explains why when the sheet's structure could not
            be confidently discovered; otherwise ``reason`` is ``None``.
        """
        header_row = self._detect_header_row(grid)
        if header_row is None:
            return None, f"no header row could be detected in '{sheet_name}'"

        data_columns = self._detect_data_columns(grid, header_row)
        if not data_columns:
            return None, f"header row in '{sheet_name}' had no populated columns"

        grouping_rows = self._detect_grouping_rows(grid, header_row)
        unit_row = self._detect_unit_row(grid, header_row, data_columns)
        data_start, data_end = self._detect_data_row_bounds(
            grid, header_row, unit_row, data_columns
        )
        if data_start is None or data_end is None:
            return None, f"no data rows could be detected in '{sheet_name}'"

        timestamp_column = self._detect_timestamp_column(
            grid, data_start, data_end, data_columns
        )

        layout = SheetLayout(
            header_row=header_row,
            grouping_rows=grouping_rows,
            unit_row=unit_row,
            data_start_row=data_start,
            data_end_row=data_end,
            timestamp_column=timestamp_column,
            data_columns=tuple(data_columns),
        )
        return layout, None

    # ------------------------------------------------------------------
    # Structural detection helpers
    # ------------------------------------------------------------------

    def _detect_header_row(self, grid: List[List[Any]]) -> Optional[int]:
        """Finds the row most likely to hold column labels.

        A header row is identified as the row, within the first
        ``max_header_search_rows`` rows, that is predominantly text and is
        immediately followed by a row that is predominantly data
        (numbers or dates) across the same columns. Among rows that pass
        this test, the one with the most populated cells is preferred,
        since a genuine header row labels every data column while a
        sparser row above it is more likely to be a section grouping row.

        Args:
            grid: The worksheet's values as a 2D list.

        Returns:
            The 1-based row index of the header row, or ``None`` if no
            row satisfies the heuristic.
        """
        search_limit = min(self.max_header_search_rows, len(grid))
        best_row: Optional[int] = None
        best_score = -1

        for row_idx in range(1, search_limit):  # leave room for a following row
            row_values = grid[row_idx - 1]
            populated = [v for v in row_values if _is_populated(v)]
            if len(populated) < 2:
                continue
            text_ratio = sum(_is_text_like(v) for v in populated) / len(populated)
            if text_ratio < 0.6:
                continue

            next_row_idx = row_idx + 1
            if next_row_idx > len(grid):
                continue
            next_values = grid[next_row_idx - 1]
            data_hits = 0
            checked = 0
            for col_idx, value in enumerate(row_values):
                if not _is_populated(value):
                    continue
                checked += 1
                next_value = (
                    next_values[col_idx] if col_idx < len(next_values) else None
                )
                if _is_data_like(next_value):
                    data_hits += 1
            if checked == 0 or (data_hits / checked) < self.min_data_fill_ratio:
                continue

            score = len(populated)
            if score > best_score:
                best_score = score
                best_row = row_idx

        return best_row

    @staticmethod
    def _detect_data_columns(grid: List[List[Any]], header_row: int) -> List[int]:
        """Finds the columns that carry a metric, based on the header row.

        Args:
            grid: The worksheet's values as a 2D list.
            header_row: The 1-based header row index.

        Returns:
            The 1-based column indices with a populated header label, in
            worksheet order.
        """
        width = _row_width(grid)
        return [
            col
            for col in range(1, width + 1)
            if _is_populated(_cell(grid, header_row, col))
        ]

    @staticmethod
    def _detect_grouping_rows(
        grid: List[List[Any]], header_row: int
    ) -> Tuple[int, ...]:
        """Finds section / sub-section label rows directly above the header.

        Grouping rows are text-only, sparser than the header row (not
        every column needs a label, since a label is meant to span
        several metric columns), and sit contiguously just above the
        header row.

        Args:
            grid: The worksheet's values as a 2D list.
            header_row: The 1-based header row index.

        Returns:
            The 1-based row indices of any grouping rows, ordered from
            outermost (furthest from the header) to innermost (closest to
            the header).
        """
        grouping_rows: List[int] = []
        row_idx = header_row - 1
        while row_idx >= 1:
            row_values = grid[row_idx - 1]
            populated = [v for v in row_values if _is_populated(v)]
            if not populated:
                break
            if any(_is_data_like(v) for v in populated):
                break
            if not all(_is_text_like(v) for v in populated):
                break
            grouping_rows.append(row_idx)
            row_idx -= 1
        grouping_rows.reverse()
        return tuple(grouping_rows)

    @staticmethod
    def _detect_unit_row(
        grid: List[List[Any]],
        header_row: int,
        data_columns: Sequence[int],
    ) -> Optional[int]:
        """Finds an optional unit row directly below the header.

        A unit row is distinguished from the start of the data block by
        being predominantly short text labels (for example ``kWh`` or
        ``bar``) rather than numbers or dates.

        Args:
            grid: The worksheet's values as a 2D list.
            header_row: The 1-based header row index.
            data_columns: The column indices that carry a metric.

        Returns:
            The 1-based row index of the unit row, or ``None`` if the row
            immediately below the header already looks like data.
        """
        candidate_row = header_row + 1
        if candidate_row > len(grid):
            return None

        values = [_cell(grid, candidate_row, col) for col in data_columns]
        populated = [v for v in values if _is_populated(v)]
        if not populated:
            return None
        if any(_is_data_like(v) for v in populated):
            return None

        short_text_ratio = sum(
            isinstance(v, str) and len(v.strip()) <= 15 for v in populated
        ) / len(populated)
        if short_text_ratio < 0.6:
            return None
        return candidate_row

    def _detect_data_row_bounds(
        self,
        grid: List[List[Any]],
        header_row: int,
        unit_row: Optional[int],
        data_columns: Sequence[int],
    ) -> Tuple[Optional[int], Optional[int]]:
        """Finds the first and last rows of the contiguous data block.

        Args:
            grid: The worksheet's values as a 2D list.
            header_row: The 1-based header row index.
            unit_row: The 1-based unit row index, if one was detected.
            data_columns: The column indices that carry a metric.

        Returns:
            A ``(first_row, last_row)`` tuple, both 1-based and inclusive,
            or ``(None, None)`` if no data rows were found.
        """
        start_row = (unit_row or header_row) + 1
        first_row: Optional[int] = None
        last_row: Optional[int] = None
        anchor_ratio: Optional[float] = None

        for row_idx in range(start_row, len(grid) + 1):
            values = [_cell(grid, row_idx, col) for col in data_columns]
            populated = [v for v in values if _is_populated(v)]
            fill_ratio = len(populated) / len(data_columns) if data_columns else 0.0

            if first_row is None:
                # The first qualifying row anchors the expected fill level
                # for the rest of the block; later rows are compared
                # against it rather than a single fixed threshold, so a
                # sudden collapse in populated columns (for example
                # leftover template rows past the real data) ends the
                # block even if it still clears the absolute minimum.
                if fill_ratio >= self.min_data_fill_ratio:
                    first_row = row_idx
                    last_row = row_idx
                    anchor_ratio = fill_ratio
                continue

            required_ratio = max(self.min_data_fill_ratio, (anchor_ratio or 0.0) * 0.5)
            if fill_ratio >= required_ratio:
                last_row = row_idx
            else:
                # A row far sparser than the established data block ends
                # the contiguous run.
                break

        return first_row, last_row

    @staticmethod
    def _detect_timestamp_column(
        grid: List[List[Any]],
        data_start: int,
        data_end: int,
        data_columns: Sequence[int],
    ) -> Optional[int]:
        """Finds the column, if any, whose values are mostly timestamps.

        Args:
            grid: The worksheet's values as a 2D list.
            data_start: The first data row.
            data_end: The last data row.
            data_columns: The column indices that carry a metric.

        Returns:
            The 1-based column index most likely to hold timestamps, or
            ``None`` if no column qualifies.
        """
        best_col: Optional[int] = None
        best_ratio = 0.0

        for col in data_columns:
            values = [_cell(grid, row, col) for row in range(data_start, data_end + 1)]
            populated = [v for v in values if _is_populated(v)]
            if not populated:
                continue
            date_hits = sum(
                isinstance(v, datetime) or _parse_datetime(v) is not None
                for v in populated
            )
            ratio = date_hits / len(populated)
            if ratio > best_ratio and ratio >= 0.8:
                best_ratio = ratio
                best_col = col

        return best_col


# ----------------------------------------------------------------------
# Model construction
# ----------------------------------------------------------------------

class SheetModelBuilder:
    """Builds Metric/Section/SubSection model objects from a known layout.

    The builder assumes structure discovery has already happened; it
    never inspects a sheet to *find* its header, grouping rows, or data
    bounds — it only reads the cells that a supplied :class:`SheetLayout`
    already points to.
    """

    def build_section(
        self,
        grid: List[List[Any]],
        layout: SheetLayout,
        sheet_name: str,
    ) -> Section:
        """Folds a sheet's grouping rows and metrics into a top-level Section.

        When two grouping rows were detected, the outer row becomes the
        top-level :class:`Section` boundary and the inner row becomes
        nested :class:`SubSection` boundaries. When one grouping row was
        detected, it becomes the sub-section boundary directly beneath a
        single sheet-level section. When none were detected, every metric
        is attached directly to a single sheet-level section.

        Args:
            grid: The worksheet's values as a 2D list.
            layout: The sheet's already-discovered structural layout.
            sheet_name: The worksheet's name, used as the section's
                fallback label.

        Returns:
            The assembled top-level :class:`Section` for this sheet.
        """
        metrics_by_column = self._build_metrics(grid, layout)
        date_range = self._compute_date_range(grid, layout)

        if not layout.grouping_rows:
            metrics = [
                metrics_by_column[c]
                for c in layout.data_columns
                if c in metrics_by_column
            ]
            return self._new_section(sheet_name, sheet_name, metrics, date_range)

        outer_row = layout.grouping_rows[0]
        inner_row = layout.grouping_rows[1] if len(layout.grouping_rows) > 1 else None

        outer_labels = self._forward_fill_labels(grid, outer_row, layout.data_columns)
        inner_labels = (
            self._forward_fill_labels(grid, inner_row, layout.data_columns)
            if inner_row is not None
            else {}
        )

        section = self._new_section(_slugify(sheet_name), sheet_name, [], date_range)
        section.subsections = self._group_into_subsections(
            layout.data_columns,
            outer_labels,
            inner_labels,
            metrics_by_column,
            date_range,
        )
        # Roll up units from every sub-section onto the parent section.
        for subsection in section.subsections:
            for unit in subsection.units:
                if unit and unit not in section.units:
                    section.units.append(unit)
        return section

    # ------------------------------------------------------------------
    # Metric construction
    # ------------------------------------------------------------------

    def _build_metrics(
        self, grid: List[List[Any]], layout: SheetLayout
    ) -> Dict[int, Metric]:
        """Builds one :class:`Metric` per data column.

        Args:
            grid: The worksheet's values as a 2D list.
            layout: The sheet's already-discovered structural layout.

        Returns:
            A mapping of column index to its constructed :class:`Metric`.
        """
        metrics: Dict[int, Metric] = {}
        timestamps = self._collect_timestamps(grid, layout)

        for col in layout.data_columns:
            if col == layout.timestamp_column:
                continue

            header_value = _cell(grid, layout.header_row, col)
            display_name = str(header_value).strip()
            name = _slugify(display_name)
            unit = self._extract_unit(grid, layout, col)

            history = self._collect_metric_history(grid, layout, col, timestamps)
            current_value, current_timestamp = self._latest_observation(history)

            metrics[col] = Metric(
                name=name,
                display_name=display_name,
                unit=unit,
                current_value=current_value,
                timestamp=current_timestamp,
                historical_values=history,
                metadata={"source_column": col, "source_row": layout.header_row},
            )

        return metrics

    @staticmethod
    def _collect_timestamps(
        grid: List[List[Any]], layout: SheetLayout
    ) -> List[Optional[datetime]]:
        """Reads the timestamp column's values across the data block.

        Args:
            grid: The worksheet's values as a 2D list.
            layout: The sheet's already-discovered structural layout.

        Returns:
            One entry per data row, in row order; ``None`` where no
            timestamp column was detected or a row's value could not be
            parsed.
        """
        if layout.timestamp_column is None:
            return [None] * (layout.data_end_row - layout.data_start_row + 1)

        timestamps: List[Optional[datetime]] = []
        for row in range(layout.data_start_row, layout.data_end_row + 1):
            raw_value = _cell(grid, row, layout.timestamp_column)
            timestamps.append(_parse_datetime(raw_value))
        return timestamps

    @staticmethod
    def _extract_unit(grid: List[List[Any]], layout: SheetLayout, col: int) -> str:
        """Reads a column's unit label from the unit row, if one exists."""
        if layout.unit_row is None:
            return ""
        value = _cell(grid, layout.unit_row, col)
        return str(value).strip() if _is_populated(value) else ""

    @staticmethod
    def _collect_metric_history(
        grid: List[List[Any]],
        layout: SheetLayout,
        col: int,
        timestamps: Sequence[Optional[datetime]],
    ) -> List[MetricDataPoint]:
        """Builds the ordered history of observations for one column.

        Args:
            grid: The worksheet's values as a 2D list.
            layout: The sheet's already-discovered structural layout.
            col: The 1-based column index of the metric.
            timestamps: The per-row timestamps aligned to the data block.

        Returns:
            One :class:`MetricDataPoint` per data row, in row order.
        """
        history: List[MetricDataPoint] = []
        for offset, row in enumerate(
            range(layout.data_start_row, layout.data_end_row + 1)
        ):
            raw_value = _cell(grid, row, col)
            numeric_value = _coerce_numeric(raw_value)
            timestamp = timestamps[offset] if offset < len(timestamps) else None
            history.append(
                MetricDataPoint(
                    timestamp=timestamp or datetime.min,
                    value=numeric_value,
                    metadata={} if timestamp else {"source_row": row},
                )
            )
        return history

    @staticmethod
    def _latest_observation(
        history: Sequence[MetricDataPoint],
    ) -> Tuple[Optional[float], Optional[datetime]]:
        """Finds the most recent non-``None`` value in a metric's history."""
        for point in reversed(history):
            if point.value is not None:
                timestamp = point.timestamp if point.timestamp != datetime.min else None
                return point.value, timestamp
        return None, None

    # ------------------------------------------------------------------
    # Section / sub-section assembly
    # ------------------------------------------------------------------

    def _group_into_subsections(
        self,
        data_columns: Sequence[int],
        outer_labels: Dict[int, str],
        inner_labels: Dict[int, str],
        metrics_by_column: Dict[int, Metric],
        date_range: Optional[DateRange],
    ) -> List[SubSection]:
        """Groups metrics into sub-sections, preserving workbook order.

        A new sub-section boundary starts whenever the combined
        (outer, inner) label pair changes from the previous data column,
        which keeps repeated department labels contiguous while still
        respecting genuinely repeated section names that occur in
        separate blocks of the sheet.

        Args:
            data_columns: The column indices carrying metrics, in order.
            outer_labels: Column-to-outer-label mapping.
            inner_labels: Column-to-inner-label mapping.
            metrics_by_column: The constructed metrics, keyed by column.
            date_range: The date range to attach to each sub-section.

        Returns:
            The ordered list of assembled :class:`SubSection` instances.
        """
        subsections: List[SubSection] = []
        current_key: Optional[Tuple[str, str]] = None
        current: Optional[SubSection] = None

        for col in data_columns:
            metric = metrics_by_column.get(col)
            if metric is None:
                continue

            outer = outer_labels.get(col, "")
            inner = inner_labels.get(col, "")
            display_name = inner or outer or metric.display_name
            key = (outer, inner)

            if key != current_key or current is None:
                current = SubSection(
                    name=_slugify(f"{outer}_{inner}" if inner else outer or display_name),
                    display_name=display_name,
                    date_range=date_range,
                    metadata={"outer_label": outer, "inner_label": inner},
                )
                subsections.append(current)
                current_key = key

            current.metrics.append(metric)
            if metric.unit and metric.unit not in current.units:
                current.units.append(metric.unit)

        return subsections

    @staticmethod
    def _forward_fill_labels(
        grid: List[List[Any]], row: int, data_columns: Sequence[int]
    ) -> Dict[int, str]:
        """Forward-fills a sparse grouping row across a set of columns.

        A grouping label typically applies to every metric column until
        the next populated label appears, since a single merged-looking
        cell spans several underlying metric columns.

        Args:
            grid: The worksheet's values as a 2D list.
            row: The 1-based grouping row index.
            data_columns: The column indices to forward-fill across.

        Returns:
            A mapping of column index to the label in effect for that
            column.
        """
        labels: Dict[int, str] = {}
        current_label = ""
        for col in data_columns:
            value = _cell(grid, row, col)
            if _is_populated(value):
                current_label = str(value).strip()
            labels[col] = current_label
        return labels

    @staticmethod
    def _new_section(
        name: str,
        display_name: str,
        metrics: List[Metric],
        date_range: Optional[DateRange],
    ) -> Section:
        """Builds a :class:`Section` with rolled-up units from its metrics."""
        section = Section(
            name=_slugify(name),
            display_name=display_name,
            metrics=metrics,
            date_range=date_range,
        )
        for metric in metrics:
            if metric.unit and metric.unit not in section.units:
                section.units.append(metric.unit)
        return section

    @staticmethod
    def _compute_date_range(
        grid: List[List[Any]], layout: SheetLayout
    ) -> Optional[DateRange]:
        """Computes the date range covered by a sheet's data block.

        Args:
            grid: The worksheet's values as a 2D list.
            layout: The sheet's already-discovered structural layout.

        Returns:
            The sheet's :class:`DateRange`, or ``None`` if no timestamp
            column was detected or no timestamp could be parsed.
        """
        if layout.timestamp_column is None:
            return None
        parsed = [
            _parse_datetime(_cell(grid, row, layout.timestamp_column))
            for row in range(layout.data_start_row, layout.data_end_row + 1)
        ]
        parsed = [p for p in parsed if p is not None]
        if not parsed:
            return None
        return DateRange(start=min(parsed), end=max(parsed))


# ----------------------------------------------------------------------
# Orchestration
# ----------------------------------------------------------------------

class ParserService:
    """Turns an already-loaded workbook into a strongly typed Workbook.

    This service performs no file I/O: it never opens a workbook from
    disk. Loading is the Loader/Repository layer's job; this service is
    handed an already-open ``openpyxl`` workbook object.

    Structure discovery can also be skipped entirely by supplying a
    pre-computed :class:`WorkbookStructure` (for example one produced
    earlier by a validator while checking the same workbook) — the
    service will reuse it instead of re-inspecting every sheet's cells.

    Attributes:
        discoverer: Performs structure discovery when no
            :class:`WorkbookStructure` is supplied to :meth:`parse`.
        model_builder: Turns a sheet's grid plus its layout into model
            objects.
    """

    def __init__(
        self,
        discoverer: Optional[WorkbookStructureDiscoverer] = None,
        model_builder: Optional[SheetModelBuilder] = None,
        max_header_search_rows: int = 30,
        min_data_fill_ratio: float = 0.3,
    ) -> None:
        """Initializes the service, optionally injecting collaborators.

        Args:
            discoverer: The structure discoverer to use. If omitted, a
                default :class:`WorkbookStructureDiscoverer` is built
                from ``max_header_search_rows`` and ``min_data_fill_ratio``.
            model_builder: The model builder to use. If omitted, a
                default :class:`SheetModelBuilder` is used.
            max_header_search_rows: Passed to the default discoverer when
                one is not supplied directly.
            min_data_fill_ratio: Passed to the default discoverer when
                one is not supplied directly.
        """
        self.discoverer = discoverer or WorkbookStructureDiscoverer(
            max_header_search_rows=max_header_search_rows,
            min_data_fill_ratio=min_data_fill_ratio,
        )
        self.model_builder = model_builder or SheetModelBuilder()

    def discover(self, raw_workbook: OpenpyxlWorkbook) -> WorkbookStructure:
        """Discovers a workbook's structure without parsing it.

        Exposed so a validator can compute and cache a
        :class:`WorkbookStructure` once, then hand it back into
        :meth:`parse` later to avoid rediscovering the same layout.

        Args:
            raw_workbook: An already-loaded ``openpyxl`` workbook.

        Returns:
            The discovered :class:`WorkbookStructure`.
        """
        return self.discoverer.discover(raw_workbook)

    def parse(
        self,
        raw_workbook: OpenpyxlWorkbook,
        workbook_name: Optional[str] = None,
        structure: Optional[WorkbookStructure] = None,
        metadata: Optional[WorkbookMetadata] = None,
    ) -> Workbook:
        """Parses an already-loaded workbook into a typed :class:`Workbook`.

        Args:
            raw_workbook: An already-open ``openpyxl`` workbook, as
                produced by the Loader/Repository layer. This method
                never opens a file itself.
            workbook_name: A human-friendly name for the workbook. If
                omitted, ``"workbook"`` is used; callers that care about
                a file-derived name should supply one (typically derived
                from the source path they used to load it).
            structure: A previously computed :class:`WorkbookStructure`
                (for example produced by a validator). When supplied,
                structure discovery is skipped entirely and this
                structure is used as-is. When omitted, the service
                discovers structure itself.
            metadata: Descriptive metadata about the workbook file. When
                omitted, minimal metadata (just a load timestamp) is
                recorded, since this service has no file-system access
                of its own.

        Returns:
            A populated :class:`Workbook`, including every section
            discovered across every sheet, its overall date range, the
            distinct units encountered, and any non-fatal warnings raised
            while parsing.
        """
        warnings: List[str] = []
        errors: List[str] = []

        workbook_model = Workbook(
            name=workbook_name or "workbook",
            metadata=metadata or WorkbookMetadata(
                source_path=None,
                file_size_bytes=None,
                loaded_at=datetime.now(),
            ),
        )
        workbook_model.available_sheets = list(raw_workbook.sheetnames)

        resolved_structure = structure or self.discoverer.discover(raw_workbook)

        for sheet_name in raw_workbook.sheetnames:
            worksheet = raw_workbook[sheet_name]

            try:
                grid = _read_grid(worksheet)
            except Exception as exc:  # noqa: BLE001 - one bad sheet shouldn't sink the rest
                warnings.append(f"Sheet '{sheet_name}' could not be read: {exc}")
                continue

            layout = resolved_structure.sheet_layouts.get(sheet_name)
            if layout is None:
                reason = resolved_structure.unparsable_sheets.get(
                    sheet_name, "sheet structure could not be discovered"
                )
                warnings.append(f"Sheet '{sheet_name}': {reason}; skipped.")
                continue

            try:
                section = self.model_builder.build_section(grid, layout, sheet_name)
            except Exception as exc:  # noqa: BLE001 - one bad sheet shouldn't sink the rest
                warnings.append(f"Sheet '{sheet_name}' could not be parsed: {exc}")
                continue

            if section is not None:
                workbook_model.sections.append(section)

        self._finalize_workbook(workbook_model, warnings, errors)
        return workbook_model

    # ------------------------------------------------------------------
    # Workbook-level assembly
    # ------------------------------------------------------------------

    def _finalize_workbook(
        self,
        workbook_model: Workbook,
        warnings: List[str],
        errors: List[str],
    ) -> None:
        """Derives workbook-wide aggregates and validation status.

        Populates the workbook's overall ``date_range`` and ``units``
        from whatever sections were discovered, and sets a
        ``validation_status`` reflecting whether any sections were found
        and whether any warnings or errors were raised while parsing.

        Args:
            workbook_model: The workbook being assembled, mutated in
                place.
            warnings: Non-fatal messages accumulated while parsing.
            errors: Fatal messages accumulated while parsing.
        """
        workbook_model.date_range = self._merge_date_ranges(
            section.date_range for section in workbook_model.sections
        )
        workbook_model.units = self._collect_units(workbook_model.sections)
        workbook_model.warnings = warnings
        workbook_model.errors = errors

        if errors:
            workbook_model.validation_status = ValidationStatus.INVALID
        elif not workbook_model.sections:
            workbook_model.validation_status = ValidationStatus.INVALID
            workbook_model.errors.append(
                "No parsable sections were discovered in this workbook."
            )
        elif warnings:
            workbook_model.validation_status = ValidationStatus.VALID_WITH_WARNINGS
        else:
            workbook_model.validation_status = ValidationStatus.VALID

    @staticmethod
    def _merge_date_ranges(ranges: Sequence[Optional[DateRange]]) -> Optional[DateRange]:
        """Combines several date ranges into their overall span."""
        starts = [r.start for r in ranges if r is not None and r.start is not None]
        ends = [r.end for r in ranges if r is not None and r.end is not None]
        if not starts and not ends:
            return None
        return DateRange(
            start=min(starts) if starts else None,
            end=max(ends) if ends else None,
        )

    @staticmethod
    def _collect_units(sections: Sequence[Section]) -> List[str]:
        """Gathers the distinct units used anywhere across all sections."""
        seen: List[str] = []
        for section in sections:
            for unit in section.units:
                if unit and unit not in seen:
                    seen.append(unit)
            for subsection in section.subsections:
                for unit in subsection.units:
                    if unit and unit not in seen:
                        seen.append(unit)
        return seen

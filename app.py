"""
app.py — INSTRUMENTED FOR RUNTIME HANG DIAGNOSIS (TEMPORARY LOGGING)

All original business logic, return values, and architecture are
unchanged. Only ENTER/EXIT/BEFORE/AFTER debug logging has been added
to identify exactly where execution stops. Remove all lines marked
"# DEBUG" once the hang is diagnosed.
"""
from __future__ import annotations

import sys
import traceback
from io import BytesIO
from types import SimpleNamespace
from typing import List, Optional

import streamlit as st
from openpyxl import load_workbook as load_openpyxl_workbook

from components.layout import configure_page, inject_global_styles, page_container
from components.theme import get_global_css

from data.github_loader import load_workbook_from_github
from data.repository import WorkbookRepository

from services.chart_service import ChartService
from services.dashboard_service import DashboardService, OverviewPageData
from services.filter_service import FilterService
from services.kpi_service import KPIService
from services.parser_service import ParserService
from services.section_service import SectionService
from services.summary_service import SummaryService
from services.workbook_service import WorkbookService

__all__ = ["main"]


# ---------------------------------------------------------------------
# DEBUG: logging helper (temporary)
# ---------------------------------------------------------------------
def _dbg(msg: str) -> None:  # DEBUG
    print(f"[DEBUG] {msg}", file=sys.stderr, flush=True)  # DEBUG


_DASHBOARD_TITLE = "Engineering Monitoring Dashboard"
_DASHBOARD_ICON = "\U0001F3ED"

_WORKBOOK_SOURCE_PATH = "github-workbook"

_SESSION_KEY_PAGE_DATA = "em_app__overview_page_data"
_SESSION_KEY_LOAD_ERROR = "em_app__load_error"

_NAV_PAGES = [
    st.Page("pages/overview.py", title="Overview", icon="🏠", url_path="overview"),
]


def _github_loader_adapter() -> object:
    """Wrap ``load_workbook_from_github`` to satisfy ``LoaderLike``."""

    def _load(source_path: str) -> SimpleNamespace:
        _dbg("ENTER app._github_loader_adapter._load")  # DEBUG
        try:
            _dbg("BEFORE data.github_loader.load_workbook_from_github()")  # DEBUG
            stream: BytesIO = load_workbook_from_github()
            _dbg("AFTER data.github_loader.load_workbook_from_github()")  # DEBUG

            try:  # DEBUG
                size_bytes = len(stream.getbuffer())  # DEBUG
                _dbg(f"Workbook download size: {size_bytes} bytes")  # DEBUG
            except Exception as size_exc:  # DEBUG
                _dbg(f"Could not determine download size: {size_exc}")  # DEBUG

            stream.seek(0)
            _dbg("BEFORE openpyxl.load_workbook(filename=stream)")  # DEBUG
            raw_workbook = load_openpyxl_workbook(filename=stream, data_only=True)
            _dbg("AFTER openpyxl.load_workbook(filename=stream)")  # DEBUG

            try:  # DEBUG
                _dbg(f"Workbook sheet names: {raw_workbook.sheetnames}")  # DEBUG
            except Exception as sheet_exc:  # DEBUG
                _dbg(f"Could not read sheet names: {sheet_exc}")  # DEBUG

            result = SimpleNamespace(
                raw_workbook=raw_workbook,
                source_path=source_path,
                metadata=None,
                success=True,
                error=None,
            )
            _dbg("EXIT app._github_loader_adapter._load (success)")  # DEBUG
            return result
        except Exception:
            _dbg("EXCEPTION in app._github_loader_adapter._load")  # DEBUG
            _dbg(traceback.format_exc())  # DEBUG
            raise

    return SimpleNamespace(load=_load)


def _passthrough_validator_adapter() -> object:
    """Stand in for ``ValidatorLike`` until a real validator exists."""

    def _validate(raw_workbook: object) -> SimpleNamespace:
        _dbg("ENTER app._passthrough_validator_adapter._validate")  # DEBUG
        try:
            result = SimpleNamespace(is_valid=True, structure=None, errors=[], warnings=[])
            _dbg(f"Validator result: is_valid={result.is_valid}, "
                 f"structure={result.structure}, errors={result.errors}, "
                 f"warnings={result.warnings}")  # DEBUG
            _dbg("EXIT app._passthrough_validator_adapter._validate")  # DEBUG
            return result
        except Exception:
            _dbg("EXCEPTION in app._passthrough_validator_adapter._validate")  # DEBUG
            _dbg(traceback.format_exc())  # DEBUG
            raise

    return SimpleNamespace(validate=_validate)


@st.cache_resource(show_spinner=False)
def _build_dashboard_service() -> DashboardService:
    """Construct the singleton `DashboardService`, wiring the full DI chain."""
    _dbg("ENTER app._build_dashboard_service")  # DEBUG
    try:
        repository = WorkbookRepository(
            loader=_github_loader_adapter(),
            validator=_passthrough_validator_adapter(),
            parser_service=ParserService(),
        )
        _dbg("Constructed WorkbookRepository")  # DEBUG

        workbook_service = WorkbookService(repository=repository)
        _dbg("Constructed WorkbookService")  # DEBUG

        section_service = SectionService(workbook_service=workbook_service)
        _dbg("Constructed SectionService")  # DEBUG

        filter_service = FilterService()
        _dbg("Constructed FilterService")  # DEBUG

        kpi_service = KPIService(section_service=section_service)
        _dbg("Constructed KPIService")  # DEBUG

        summary_service = SummaryService(
            workbook_service=workbook_service,
            section_service=section_service,
            filter_service=filter_service,
            kpi_service=kpi_service,
        )
        _dbg("Constructed SummaryService")  # DEBUG

        chart_service = ChartService(
            section_service=section_service,
            filter_service=filter_service,
        )
        _dbg("Constructed ChartService")  # DEBUG

        service = DashboardService(
            workbook_service=workbook_service,
            section_service=section_service,
            filter_service=filter_service,
            kpi_service=kpi_service,
            summary_service=summary_service,
            chart_service=chart_service,
        )
        _dbg("Constructed DashboardService")  # DEBUG
        _dbg("EXIT app._build_dashboard_service (success)")  # DEBUG
        return service
    except Exception:
        _dbg("EXCEPTION in app._build_dashboard_service")  # DEBUG
        _dbg(traceback.format_exc())  # DEBUG
        raise


def _load_overview_page_data(
    dashboard_service: DashboardService,
) -> Optional[OverviewPageData]:
    """Load the Overview page's data via ``DashboardService``, once per session."""
    _dbg("ENTER app._load_overview_page_data")  # DEBUG
    try:
        _dbg("BEFORE dashboard_service.get_overview_page_data()")  # DEBUG
        page_data = dashboard_service.get_overview_page_data(
            source_path=_WORKBOOK_SOURCE_PATH
        )
        _dbg("AFTER dashboard_service.get_overview_page_data()")  # DEBUG
        _dbg(f"OverviewPageData created: header={page_data.header}, "
             f"top_kpi_cards_count={len(page_data.top_kpi_cards)}, "
             f"summary_cards_count={len(page_data.summary_cards)}, "
             f"expandable_sections_count={len(page_data.expandable_sections)}")  # DEBUG
    except Exception as exc:  # noqa: BLE001 - surfaced to the user below
        _dbg("EXCEPTION in app._load_overview_page_data")  # DEBUG
        _dbg(traceback.format_exc())  # DEBUG
        st.session_state[_SESSION_KEY_LOAD_ERROR] = exc
        st.session_state[_SESSION_KEY_PAGE_DATA] = None
        _dbg("EXIT app._load_overview_page_data (failure)")  # DEBUG
        return None

    st.session_state[_SESSION_KEY_LOAD_ERROR] = None
    st.session_state[_SESSION_KEY_PAGE_DATA] = page_data
    _dbg("EXIT app._load_overview_page_data (success)")  # DEBUG
    return page_data


def _register_navigation(pages: List[st.Page]) -> None:
    """Register the application's multipage navigation."""
    _dbg("ENTER app._register_navigation")  # DEBUG
    try:
        st.navigation(pages).run()
        _dbg("EXIT app._register_navigation")  # DEBUG
    except Exception:
        _dbg("EXCEPTION in app._register_navigation")  # DEBUG
        _dbg(traceback.format_exc())  # DEBUG
        raise


def main() -> None:
    """Application entry point: configure, wire, load, and launch."""
    _dbg("ENTER app.main")  # DEBUG
    try:
        configure_page(_DASHBOARD_TITLE, _DASHBOARD_ICON)
        _dbg("configure_page done")  # DEBUG
        inject_global_styles(get_global_css())
        _dbg("inject_global_styles done")  # DEBUG

        if "dashboard_service" not in st.session_state:
            _dbg("BEFORE app._build_dashboard_service()")  # DEBUG
            st.session_state["dashboard_service"] = _build_dashboard_service()
            _dbg("AFTER app._build_dashboard_service()")  # DEBUG
        else:
            _dbg("dashboard_service already in session_state; skipping build")  # DEBUG

        if "workbook_source_path" not in st.session_state:
            st.session_state["workbook_source_path"] = _WORKBOOK_SOURCE_PATH

        dashboard_service: DashboardService = st.session_state["dashboard_service"]

        if _SESSION_KEY_PAGE_DATA not in st.session_state:
            with st.spinner("Loading workbook..."):
                st.info("🔄 Connecting to workbook source...")
                _dbg("BEFORE app._load_overview_page_data()")  # DEBUG
                _load_overview_page_data(dashboard_service)
                _dbg("AFTER app._load_overview_page_data()")  # DEBUG
        else:
            _dbg(f"{_SESSION_KEY_PAGE_DATA} already in session_state; skipping load")  # DEBUG

        load_error = st.session_state.get(_SESSION_KEY_LOAD_ERROR)
        if load_error is not None:
            _dbg(f"Load error present: {load_error}")  # DEBUG
            st.error(f"Failed to load workbook:\n\n{load_error}")
            if st.button("Retry"):
                st.session_state[_SESSION_KEY_PAGE_DATA] = None
                st.session_state[_SESSION_KEY_LOAD_ERROR] = None
                st.rerun()
            _dbg("EXIT app.main (returned early due to load error)")  # DEBUG
            return

        with page_container():
            _dbg("BEFORE app._register_navigation()")  # DEBUG
            _register_navigation(_NAV_PAGES)
            _dbg("AFTER app._register_navigation()")  # DEBUG

        _dbg("EXIT app.main (success)")  # DEBUG
    except Exception:
        _dbg("EXCEPTION in app.main")  # DEBUG
        _dbg(traceback.format_exc())  # DEBUG
        raise


if __name__ == "__main__":
    main()
